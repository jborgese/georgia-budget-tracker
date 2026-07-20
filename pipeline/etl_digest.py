"""Transform DOR tax digest exports into millage-rate datasets.

The Georgia Department of Revenue's consolidated tax digest, republished as
annual Excel exports by UGA's Carl Vinson Institute (georgiadata.org, source
ids ``dor_digest_YYYY`` in pipeline/sources.json), carries one row per
county and taxing district: parcel counts, assessed values, exemptions, and
the millage rates and levies for maintenance & operations (M&O) and bond
debt. This is the "what a resident is charged" counterpart to the RLGF
"what governments spend" filings, and the only machine-readable view of
sub-county taxing districts (fire districts, CIDs, school bond levies).

Layout notes the parser relies on:
- Sheet names vary by vintage ("2016 digest", "Job", "Digest Export.4.1.25")
  but the first sheet always carries the data.
- Vintages 1990-2013 use raw DOR field names ("txpyr-name", "tax-mlg-mo");
  2014+ use display names ("County Name", "Millage Rate-M&O"). Both name
  sets appear in DIGEST_COLUMNS and carry identical semantics — the levy
  identity (tax = assessed x millage / 1000) holds row-for-row in both.
- District Code 0 is the county aggregate row (parcels and assessed values
  for the whole county, no millage of its own); the STATE district levied
  0.25 mills through 2015 and 0 from 2016 on (the state millage was phased
  out after 2015).
- An empty millage cell means the district did not report a rate — kept as
  null, never coerced to 0. A literal 0 is a real levied rate of zero.
- The 2023 and 2024 vintages key some rows under truncated or suffixed
  county names ("CHARLT TB", "GORDCO"); canonical_county resolves them
  against the Census roster by progressively dropping trailing tokens.
- A district split across rows with one rate (2022 CHEROKEE / HOLLY SPRINGS
  FIRE) is collapsed by summing its amounts; conflicting rates still fail.
- KNOWN_MISSING documents counties absent from a vintage upstream (WAYNE
  has no rows in the 2014-2015 exports, FULTON none in 2017-2018); any
  other gap, or a documented gap that heals, fails loudly so the registry
  stays deliberate.
- DOR itself directs readers to county tax commissioners for authoritative
  figures; these exports are the state's compiled view.

Outputs:
- data/processed/digest.parquet          one row per (county, district, year),
  full history (1990+)
- data/processed/counties/millage.json   per-county districts with rates and
  levies per tax year, plus the county aggregate assessed values — windowed
  to the most recent MILLAGE_WEB_YEARS tax years so county-page payloads
  stay small (the detailed rates table renders the latest year; the parquet
  keeps the full series)
- data/processed/counties/millage_history.json   rates-only [M&O, bond]
  pairs per district and tax year across the full span, plus KNOWN_MISSING
  inverted per county slug, so the county pages' rate-history view and its
  gap disclosure derive from the same registry the parser enforces

Usage: etl_digest.py [dor_digest_YYYY ...]
With no arguments every year is refreshed; with source-id arguments only
those files are re-downloaded (other years reuse the raw file on disk,
downloading it first if absent).
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

import runlog
from fetching import download_file_stdlib

ROOT = Path(__file__).resolve().parent.parent
SOURCES_FILE = ROOT / "pipeline" / "sources.json"
COUNTY_ROSTER = set(json.loads(
    (Path(__file__).resolve().parent / "ga_counties.json").read_text()))
RAW_DIR = ROOT / "data" / "raw"
PARQUET_FILE = ROOT / "data" / "processed" / "digest.parquet"
MILLAGE_FILE = ROOT / "data" / "processed" / "counties" / "millage.json"
MILLAGE_HISTORY_FILE = (ROOT / "data" / "processed" / "counties"
                        / "millage_history.json")
SOURCE_PREFIX = "dor_digest_"
REQUEST_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (compatible; georgia-budget-tracker/0.1; "
                   "+https://github.com/jborgese/georgia-budget-tracker)"),
}
COUNTY_TOTAL_CODE = 0
MILLAGE_WEB_YEARS = 10
KNOWN_MISSING = {
    2014: frozenset({"WAYNE"}),
    2015: frozenset({"WAYNE"}),
    2017: frozenset({"FULTON"}),
    2018: frozenset({"FULTON"}),
}

DIGEST_COLUMNS = {
    "county": ("County Name", "txpyr-name"),
    "district": ("District Name", "txpyr-dist-name"),
    "district_code": ("District Code", "txpyr-dist-did"),
    "tax_year": ("Tax Year", "rtn-period-taxYr"),
    "parcels": ("Total Parcels", "pCnt"),
    "assessed_mo": ("Total Assessed Value-M&O", "tax-valAmt-mo"),
    "assessed_bond": ("Total Assessed Value-Bond", "tax-valAmt-bnd"),
    "millage_mo": ("Millage Rate-M&O", "tax-mlg-mo"),
    "millage_bond": ("Millage Rate-Bond", "tax-mlg-bnd"),
    "tax_mo": ("Total Tax-M&O", "tax-taxAmt-mo"),
    "tax_bond": ("Total Tax-Bond", "tax-taxAmt-bnd"),
}
TEXT_FIELDS = {"county", "district"}
RATE_FIELDS = {"millage_mo", "millage_bond"}
AMOUNT_FIELDS = {"parcels", "assessed_mo", "assessed_bond",
                 "tax_mo", "tax_bond"}
DISTRICT_KEYS = ["county", "district_code", "district"]


def collapse_split_districts(frame: pd.DataFrame, name: str) -> pd.DataFrame:
    if not frame.duplicated(DISTRICT_KEYS).any():
        return frame
    rates = frame.groupby(DISTRICT_KEYS)[list(RATE_FIELDS)].nunique()
    conflicted = rates[rates.gt(1).any(axis=1)]
    if not conflicted.empty:
        raise SystemExit(
            f"{name} carries duplicate district rows with conflicting "
            f"rates: {sorted({county for county, _, _ in conflicted.index})[:5]}.")
    return frame.groupby(DISTRICT_KEYS, as_index=False).agg({
        field: (lambda s: s.sum(min_count=1)) if field in AMOUNT_FIELDS
        else "first"
        for field in frame.columns if field not in DISTRICT_KEYS})


def digest_sources() -> dict[str, dict]:
    sources = json.loads(SOURCES_FILE.read_text())["sources"]
    return {s["id"]: s for s in sources if s["id"].startswith(SOURCE_PREFIX)}


def raw_file(source_id: str) -> Path:
    return RAW_DIR / f"{source_id}.xlsx"


def source_year(source_id: str) -> int:
    return int(source_id.removeprefix(SOURCE_PREFIX))


def unique_prefix_match(compressed: str) -> str | None:
    matches = [county for county in COUNTY_ROSTER
               if compressed and county.replace(" ", "").startswith(compressed)]
    return matches[0] if len(matches) == 1 else None


def roster_match(candidate: str) -> str | None:
    if candidate in COUNTY_ROSTER:
        return candidate
    trimmed = candidate.removesuffix(" SO").removesuffix(" COUNTY").strip()
    if trimmed in COUNTY_ROSTER:
        return trimmed
    compressed = trimmed.replace(" ", "")
    return (unique_prefix_match(compressed)
            or (unique_prefix_match(compressed[:-2])
                if compressed.endswith("CO") else None))


def canonical_county(name: str) -> str:
    tokens = name.strip().upper().split()
    for length in range(len(tokens), 0, -1):
        match = roster_match(" ".join(tokens[:length]))
        if match:
            return match
    raise SystemExit(f"County name {name!r} does not resolve to the Census "
                     "roster.")


def county_slug(county: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", county.lower()).strip("-")


def parse_year(path: Path, tax_year: int) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    missing = [names[0] for names in DIGEST_COLUMNS.values()
               if not any(name in header for name in names)]
    if missing:
        raise SystemExit(
            f"{path.name} lacks expected digest columns {missing} — "
            "layout may have changed.")
    positions = {field: next(header.index(name) for name in names
                             if name in header)
                 for field, names in DIGEST_COLUMNS.items()}
    records = []
    for cells in rows:
        county = cells[positions["county"]]
        if county is None or not str(county).strip():
            continue
        record = {field: cells[position]
                  for field, position in positions.items()}
        for field in TEXT_FIELDS:
            record[field] = str(record[field]).strip().upper()
        record["county"] = canonical_county(record["county"])
        records.append(record)
    workbook.close()
    frame = pd.DataFrame.from_records(records)
    parsed_years = set(frame.tax_year.dropna().astype(int).unique())
    if parsed_years != {tax_year}:
        raise SystemExit(
            f"{path.name} carries tax years {sorted(parsed_years)}, "
            f"expected {tax_year}.")
    missing = COUNTY_ROSTER - set(frame.county.unique())
    if missing != KNOWN_MISSING.get(tax_year, frozenset()):
        raise SystemExit(
            f"{path.name} is missing counties {sorted(missing)}; documented "
            f"gaps for {tax_year}: "
            f"{sorted(KNOWN_MISSING.get(tax_year, frozenset()))}.")
    return collapse_split_districts(frame, path.name).assign(
        tax_year=tax_year)


def write_parquet(frame: pd.DataFrame) -> None:
    PARQUET_FILE.parent.mkdir(parents=True, exist_ok=True)
    connection = duckdb.connect()
    connection.register("digest", frame)
    connection.execute(
        f"""
        COPY (SELECT * FROM digest
              ORDER BY county, district_code, district, tax_year)
        TO '{PARQUET_FILE}' (FORMAT PARQUET, COMPRESSION ZSTD)
        """
    )


def number(value) -> float | int | None:
    if value is None or pd.isna(value):
        return None
    numeric = float(value)
    return int(numeric) if numeric.is_integer() else round(numeric, 3)


def district_years(rows: pd.DataFrame) -> dict:
    return {
        str(int(row.tax_year)): {
            "millage_mo": number(row.millage_mo),
            "millage_bond": number(row.millage_bond),
            "tax_mo": number(row.tax_mo),
            "tax_bond": number(row.tax_bond),
        }
        for row in rows.sort_values("tax_year").itertuples(index=False)
    }


def county_total_years(rows: pd.DataFrame) -> dict:
    return {
        str(int(row.tax_year)): {
            "parcels": number(row.parcels),
            "assessed_mo": number(row.assessed_mo),
            "assessed_bond": number(row.assessed_bond),
        }
        for row in rows.sort_values("tax_year").itertuples(index=False)
    }


def county_entry(rows: pd.DataFrame) -> dict:
    totals = rows[rows.district_code == COUNTY_TOTAL_CODE]
    districts = rows[rows.district_code != COUNTY_TOTAL_CODE]
    return {
        "county": rows.county.iloc[0],
        "county_total": county_total_years(totals),
        "districts": [
            {
                "district": district,
                "code": int(district_rows.district_code.iloc[0]),
                "years": district_years(district_rows),
            }
            for (code, district), district_rows in sorted(
                districts.groupby(["district_code", "district"]),
                key=lambda item: item[0])
        ],
    }


def write_millage_json(frame: pd.DataFrame, source_ids: list[str]) -> int:
    web_years = sorted(int(y) for y in frame.tax_year.unique())[
        -MILLAGE_WEB_YEARS:]
    windowed = frame[frame.tax_year.isin(web_years)]
    counties = {
        county_slug(county): county_entry(rows)
        for county, rows in windowed.groupby("county")
    }
    document = {
        "sources": [source_id for source_id in source_ids
                    if source_year(source_id) in set(web_years)],
        "tax_years": web_years,
        "counties": dict(sorted(counties.items())),
    }
    MILLAGE_FILE.parent.mkdir(parents=True, exist_ok=True)
    MILLAGE_FILE.write_text(json.dumps(document, indent=1) + "\n")
    return len(counties)


def district_rates(rows: pd.DataFrame) -> dict:
    return {
        str(int(row.tax_year)): [number(row.millage_mo),
                                 number(row.millage_bond)]
        for row in rows.sort_values("tax_year").itertuples(index=False)
    }


def county_history_entry(rows: pd.DataFrame) -> dict:
    districts = rows[rows.district_code != COUNTY_TOTAL_CODE]
    return {
        "county": rows.county.iloc[0],
        "districts": [
            {
                "district": district,
                "code": int(code),
                "rates": district_rates(district_rows),
            }
            for (code, district), district_rows in sorted(
                districts.groupby(["district_code", "district"]),
                key=lambda item: item[0])
        ],
    }


def missing_by_county() -> dict[str, list[int]]:
    missing: dict[str, list[int]] = {}
    for year, counties in sorted(KNOWN_MISSING.items()):
        for county in sorted(counties):
            missing.setdefault(county_slug(county), []).append(year)
    return dict(sorted(missing.items()))


def write_millage_history_json(frame: pd.DataFrame,
                               source_ids: list[str]) -> None:
    document = {
        "sources": source_ids,
        "tax_years": sorted(int(y) for y in frame.tax_year.unique()),
        "known_missing": missing_by_county(),
        "counties": dict(sorted(
            (county_slug(county), county_history_entry(rows))
            for county, rows in frame.groupby("county"))),
    }
    MILLAGE_HISTORY_FILE.parent.mkdir(parents=True, exist_ok=True)
    MILLAGE_HISTORY_FILE.write_text(json.dumps(document, indent=1) + "\n")


def ensure_raw(source: dict, refresh: bool) -> Path:
    target = raw_file(source["id"])
    if refresh or not target.exists():
        try:
            download_file_stdlib(source["url"], target, source=source["id"],
                                 headers=REQUEST_HEADERS)
            print(f"Downloaded {source['url']} -> {target} "
                  f"({target.stat().st_size:,} bytes)")
        except (Exception, SystemExit):
            if not target.exists():
                raise
            runlog.log_event("fallback", source["id"],
                             reason="using committed raw file")
    return target


def main() -> int:
    requested = set(sys.argv[1:])
    sources = digest_sources()
    refresh = set(sources) & requested if requested else set(sources)
    if requested and not refresh:
        print(f"No digest sources among {sorted(requested)}; nothing to do.")
        return 0

    frames = []
    any_failed = False
    for source_id, source in sorted(sources.items()):
        try:
            path = ensure_raw(source, refresh=source_id in refresh)
            frames.append(parse_year(path, source_year(source_id)))
        except (Exception, SystemExit) as exc:
            failures = runlog.record_outcome(source_id, ok=False,
                                             error=str(exc))
            runlog.log_event("transform_failed", source_id,
                             consecutive_failures=failures,
                             error=str(exc)[:300])
            print(f"ERROR {source_id}: {exc}", file=sys.stderr)
            any_failed = True
            continue
        if source_id in refresh:
            runlog.record_outcome(source_id, ok=True)

    if not frames:
        print("No digest files parsed; existing outputs left untouched.",
              file=sys.stderr)
        return 1

    combined = pd.concat(frames, ignore_index=True)
    write_parquet(combined)
    county_count = write_millage_json(combined, sorted(sources))
    write_millage_history_json(combined, sorted(sources))
    runlog.log_event("transformed", "dor_digest", records=len(combined),
                     counties=county_count)
    print(
        f"Wrote {len(combined):,} district-year rows for "
        f"{county_count} counties "
        f"({combined.tax_year.min()}-{combined.tax_year.max()}) to "
        f"{PARQUET_FILE.relative_to(ROOT)}, "
        f"{MILLAGE_FILE.relative_to(ROOT)}, and "
        f"{MILLAGE_HISTORY_FILE.relative_to(ROOT)}"
    )
    return 1 if any_failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
