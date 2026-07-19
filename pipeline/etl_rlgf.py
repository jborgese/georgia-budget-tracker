"""Transform the TED RLGF workbooks into normalized datasets.

The UGA Carl Vinson Institute's Tax & Expenditure Data Center publishes
prebuilt Excel exports of Report of Local Government Finances data, one per
government type (source ids ``ted_rlgf_county_workbook``,
``ted_rlgf_city_workbook``, and ``ted_rlgf_consolidated_workbook`` in
pipeline/sources.json). Each sheet is one government — 152 county sheets, 522
city sheets, and 8 consolidated city-county sheets — rows are the RLGF
classification hierarchy indented five spaces per level, and columns are
fiscal years (2016 onward).

Sheet layout conventions this parser relies on (identical across the three
workbooks):
- Row 1 is ``Classification`` followed by fiscal-year columns.
- Unindented Title Case rows with no amounts ("Revenues", "Debt", ...) are
  section headers; unindented ALL CAPS rows ("TOTAL REVENUES", ...) are that
  section's total lines.
- The Revenues, Operating Expenditures, Capital Expenditures, and Debt
  (PART XI: outstanding, issued, retired, interest by debt type) sections
  are extracted; the remaining sections (cash, fund equity, personnel) are
  out of scope for now.

Outputs per government type:
- data/raw/<source_id>.xlsx                     workbook as downloaded
- data/processed/rlgf_<type>_finances.parquet   one row per
  (government, fiscal_year, classification line)
- data/processed/<dir>/<slug>.json              per-government totals plus a
  shallow (depth <= 2) breakdown for the frontend
- data/processed/<dir>/index.json               government list with latest totals
- data/processed/<dir>/sales_tax.json           itemized sales-tax revenue lines
  (LOST/SPLOST/TSPLOST/...) per government, entities with any nonzero year

The county outputs keep their original column and key names (``county``,
``counties``) so historical data files stay byte-stable; city and
consolidated outputs use the generic ``entity`` naming.

Usage: etl_rlgf.py [county|city|consolidated] [path-to-local-workbook]
The government type defaults to county. When a local path is given it is
copied into data/raw/ instead of downloading.
"""

from __future__ import annotations

import json
import re
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

import runlog
from fetching import download_file

ROOT = Path(__file__).resolve().parent.parent
SOURCES_FILE = ROOT / "pipeline" / "sources.json"
PROCESSED_DIR = ROOT / "data" / "processed"
USER_AGENT = (
    "georgia-budget-tracker/0.1 (+https://github.com/<owner>/georgia-budget-tracker; "
    "civic data project)"
)
TIMEOUT_SECONDS = 120
INDENT_WIDTH = 5
SECTIONS = {
    "Revenues": ("revenue", "revenues"),
    "Operating Expenditures": ("expenditure", "operating"),
    "Capital Expenditures": ("expenditure", "capital"),
    "Debt": ("debt", "debt"),
}
JSON_BREAKDOWN_MAX_DEPTH = 2
SALES_TAX_PATTERN = "SPLOST|Sales Tax|Sales and Use"
CONSOLIDATED_SUFFIX = re.compile(r"\s+COUNTY$")
PRESERVED_JSON = ("metrics.json", "categories.json", "index.json")


@dataclass(frozen=True)
class GovernmentType:
    key: str
    source_id: str
    entity_column: str
    json_dir: str
    index_key: str


GOVERNMENT_TYPES = {
    "county": GovernmentType("county", "ted_rlgf_county_workbook",
                             "county", "counties", "counties"),
    "city": GovernmentType("city", "ted_rlgf_city_workbook",
                           "entity", "cities", "entities"),
    "consolidated": GovernmentType("consolidated", "ted_rlgf_consolidated_workbook",
                                   "entity", "consolidated", "entities"),
}


def raw_file(government: GovernmentType) -> Path:
    return ROOT / "data" / "raw" / f"{government.source_id}.xlsx"


def parquet_file(government: GovernmentType) -> Path:
    return PROCESSED_DIR / f"rlgf_{government.key}_finances.parquet"


def json_dir(government: GovernmentType) -> Path:
    return PROCESSED_DIR / government.json_dir


def source_url(source_id: str) -> str:
    sources = json.loads(SOURCES_FILE.read_text())["sources"]
    matches = [s["url"] for s in sources if s["id"] == source_id]
    if not matches:
        raise SystemExit(f"Source {source_id!r} not found in {SOURCES_FILE}")
    return matches[0]


def canonical_entity(sheet_name: str) -> str:
    return CONSOLIDATED_SUFFIX.sub("", sheet_name.strip().upper())


def indent_depth(label: str) -> int:
    return (len(label) - len(label.lstrip(" "))) // INDENT_WIDTH


def is_section_header(name: str, depth: int, amounts: list) -> bool:
    return depth == 0 and name != name.upper() and all(a is None for a in amounts)


def to_amount(value) -> int | None:
    numeric = pd.to_numeric(value, errors="coerce")
    return None if pd.isna(numeric) else int(numeric)


def sheet_records(entity: str, header: tuple, rows: list[tuple],
                  entity_column: str = "county") -> list[dict]:
    years = [int(cell) for cell in header[1:] if cell is not None]
    section = None
    stack: list[str] = []
    records = []
    for line, cells in enumerate(rows, start=2):
        raw = cells[0]
        if raw is None or not str(raw).strip():
            continue
        label = str(raw)
        name = label.strip()
        depth = indent_depth(label)
        amounts = [to_amount(cell) for cell in cells[1 : len(years) + 1]]
        if is_section_header(name, depth, amounts):
            section = SECTIONS.get(name)
            stack = []
            continue
        if section is None:
            continue
        category, section_name = section
        stack = stack[:depth] + [name]
        path = " > ".join(stack)
        records.extend(
            {
                entity_column: entity,
                "fiscal_year": year,
                "category": category,
                "section": section_name,
                "classification": name,
                "depth": depth,
                "line": line,
                "path": path,
                "amount": amount,
            }
            for year, amount in zip(years, amounts)
            if amount is not None
        )
    return records


def parse_workbook(path: Path, entity_column: str = "county") -> pd.DataFrame:
    workbook = openpyxl.load_workbook(path, read_only=True)
    records = [
        record
        for sheet in workbook.sheetnames
        for rows in [list(workbook[sheet].iter_rows(values_only=True))]
        for record in sheet_records(canonical_entity(sheet), rows[0], rows[1:],
                                    entity_column)
    ]
    workbook.close()
    return pd.DataFrame.from_records(records)


def entity_slug(entity: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", entity.lower()).strip("-")


def write_parquet(connection: duckdb.DuckDBPyConnection,
                  government: GovernmentType) -> None:
    target = parquet_file(government)
    target.parent.mkdir(parents=True, exist_ok=True)
    connection.execute(
        f"""
        COPY (SELECT * FROM records
              ORDER BY {government.entity_column}, line, fiscal_year)
        TO '{target}' (FORMAT PARQUET, COMPRESSION ZSTD)
        """
    )


def entity_totals(connection: duckdb.DuckDBPyConnection,
                  government: GovernmentType) -> pd.DataFrame:
    return connection.execute(
        f"""
        SELECT
            {government.entity_column} AS entity,
            fiscal_year,
            sum(amount) FILTER (category = 'revenue') AS revenue,
            sum(amount) FILTER (category = 'expenditure') AS expenditure,
            sum(amount) FILTER (section = 'operating') AS expenditure_operating,
            sum(amount) FILTER (section = 'capital') AS expenditure_capital
        FROM records
        WHERE depth = 0
        GROUP BY entity, fiscal_year
        ORDER BY entity, fiscal_year
        """
    ).df()


def entity_breakdown(connection: duckdb.DuckDBPyConnection,
                     government: GovernmentType) -> pd.DataFrame:
    return connection.execute(
        f"""
        SELECT {government.entity_column} AS entity,
               classification, category, section, depth, line, path,
               map_from_entries(
                   list((fiscal_year, amount) ORDER BY fiscal_year)
               ) AS amounts
        FROM records
        WHERE depth <= ?
        GROUP BY entity, classification, category, section, depth, line, path
        ORDER BY entity, line
        """,
        [JSON_BREAKDOWN_MAX_DEPTH],
    ).df()


def totals_entry(row: pd.Series) -> dict:
    return {
        "fiscal_year": int(row.fiscal_year),
        "revenue": None if pd.isna(row.revenue) else int(row.revenue),
        "expenditure": None if pd.isna(row.expenditure) else int(row.expenditure),
        "expenditure_operating": (
            None if pd.isna(row.expenditure_operating) else int(row.expenditure_operating)
        ),
        "expenditure_capital": (
            None if pd.isna(row.expenditure_capital) else int(row.expenditure_capital)
        ),
    }


def breakdown_entry(row: pd.Series) -> dict:
    return {
        "classification": row.classification,
        "category": row.category,
        "section": row.section,
        "depth": int(row.depth),
        "path": row.path,
        "amounts": {str(year): int(amount) for year, amount in row.amounts.items()},
    }


def entity_document(government: GovernmentType, entity: str,
                    totals: pd.DataFrame, breakdown: pd.DataFrame) -> dict:
    return {
        government.entity_column: entity,
        "source": government.source_id,
        "fiscal_years": [int(year) for year in sorted(totals.fiscal_year)],
        "totals": [totals_entry(row) for row in totals.itertuples(index=False)],
        "breakdown": [breakdown_entry(row) for row in breakdown.itertuples(index=False)],
    }


def index_document(government: GovernmentType, totals: pd.DataFrame) -> dict:
    latest = totals.sort_values("fiscal_year").groupby("entity").last().reset_index()
    return {
        "source": government.source_id,
        "fiscal_years": [int(year) for year in sorted(totals.fiscal_year.unique())],
        government.index_key: [
            {
                government.entity_column: row.entity,
                "slug": entity_slug(row.entity),
                "latest_fiscal_year": int(row.fiscal_year),
                "revenue": None if pd.isna(row.revenue) else int(row.revenue),
                "expenditure": None if pd.isna(row.expenditure) else int(row.expenditure),
            }
            for row in latest.sort_values("entity").itertuples(index=False)
        ],
    }


def write_json(path: Path, document: dict) -> None:
    path.write_text(json.dumps(document, indent=1) + "\n")


def write_entity_json(connection: duckdb.DuckDBPyConnection,
                      government: GovernmentType) -> int:
    totals = entity_totals(connection, government)
    breakdown = entity_breakdown(connection, government)
    target = json_dir(government)
    target.mkdir(parents=True, exist_ok=True)
    for stale in target.glob("*.json"):
        if stale.name not in PRESERVED_JSON:
            stale.unlink()
    for entity, entity_rows in totals.groupby("entity"):
        document = entity_document(
            government, entity, entity_rows, breakdown[breakdown.entity == entity]
        )
        write_json(target / f"{entity_slug(entity)}.json", document)
    write_json(target / "index.json", index_document(government, totals))
    return totals.entity.nunique()


def sales_tax_lines(connection: duckdb.DuckDBPyConnection,
                    government: GovernmentType) -> pd.DataFrame:
    return connection.execute(
        f"""
        SELECT {government.entity_column} AS entity, classification,
               map_from_entries(
                   list((fiscal_year, amount) ORDER BY fiscal_year)
               ) AS amounts,
               min(line) AS first_line
        FROM records
        WHERE depth = 4
          AND section = 'revenues'
          AND path LIKE '%PART I TAX REVENUES%Section B - General Sales and Use Taxes%'
          AND regexp_matches(classification, ?)
        GROUP BY entity, classification
        HAVING sum(abs(amount)) > 0
        ORDER BY entity, first_line
        """,
        [SALES_TAX_PATTERN],
    ).df()


def write_sales_tax_json(connection: duckdb.DuckDBPyConnection,
                         government: GovernmentType,
                         fiscal_years: list[int]) -> None:
    lines = sales_tax_lines(connection, government)
    entities: dict[str, dict] = {}
    for entity, entity_rows in lines.groupby("entity"):
        entities[entity_slug(entity)] = {
            "entity": entity,
            "lines": [
                {
                    "classification": row.classification,
                    "amounts": {str(year): int(amount)
                                for year, amount in row.amounts.items()},
                }
                for row in entity_rows.itertuples(index=False)
            ],
        }
    document = {
        "source": government.source_id,
        "fiscal_years": fiscal_years,
        "entities": dict(sorted(entities.items())),
    }
    write_json(json_dir(government) / "sales_tax.json", document)


def parse_args(argv: list[str]) -> tuple[GovernmentType, str | None]:
    if argv and argv[0] in GOVERNMENT_TYPES:
        return GOVERNMENT_TYPES[argv[0]], argv[1] if len(argv) > 1 else None
    return GOVERNMENT_TYPES["county"], argv[0] if argv else None


def run(government: GovernmentType, local_path: str | None) -> None:
    target = raw_file(government)
    if local_path:
        target.parent.mkdir(parents=True, exist_ok=True)
        if Path(local_path).resolve() != target.resolve():
            shutil.copyfile(local_path, target)
            print(f"Copied local workbook {local_path} -> {target}")
    else:
        url = source_url(government.source_id)
        download_file(url, target, source=government.source_id)
        print(f"Downloaded {url} -> {target} ({target.stat().st_size:,} bytes)")

    records = parse_workbook(target, government.entity_column)
    if records.empty:
        raise SystemExit("Workbook parsed to zero records — layout may have changed.")

    connection = duckdb.connect()
    connection.register("records", records)
    write_parquet(connection, government)
    entity_count = write_entity_json(connection, government)
    write_sales_tax_json(connection, government,
                         sorted(int(y) for y in records.fiscal_year.unique()))

    runlog.log_event("transformed", government.source_id, records=len(records),
                     entities=entity_count)
    print(
        f"Wrote {len(records):,} records for {entity_count} "
        f"{government.key} governments "
        f"({records.fiscal_year.min()}-{records.fiscal_year.max()}) to "
        f"{parquet_file(government).relative_to(ROOT)} and "
        f"{json_dir(government).relative_to(ROOT)}/"
    )


def main() -> int:
    government, local_path = parse_args(sys.argv[1:])
    try:
        run(government, local_path)
    except (Exception, SystemExit) as exc:
        failures = runlog.record_outcome(government.source_id, ok=False, error=str(exc))
        runlog.log_event("transform_failed", government.source_id,
                         consecutive_failures=failures, error=str(exc)[:300])
        print(f"ERROR {government.source_id}: {exc}", file=sys.stderr)
        return 1
    runlog.record_outcome(government.source_id, ok=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
