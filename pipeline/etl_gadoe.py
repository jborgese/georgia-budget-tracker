"""Transform GaDOE current-year school revenues into an F-33 supplement.

The Census F-33 survey (etl_schools.py) publishes roughly 18 months after
each fiscal year closes. GaDOE's Financial Data Collection System (source id
``gadoe_revenues``) already carries the newest year, but only as headline
figures: the School System Revenues report gives Local/State/Federal/Total
revenues and the QBE FTE count per system — no ESPLOST or property-tax
detail — and its revenue basis excludes debt-service and capital-projects
funds, so totals run below the all-funds F-33 figures. This transform
therefore fetches only fiscal years newer than the latest F-33 vintage and
publishes them as an overlay next to the F-33 outputs; the F-33 series
remains the source of detail and history.

Downloads replay the portal's session-bound ASP.NET/SSRS export flow via
pipeline/gadoe.py (see that module for why header fingerprints don't work).
GaDOE systems are joined to F-33 districts by name: three-digit system codes
are Georgia's regular districts and must map — via ``match_key`` token
normalization, ``GADOE_ALIASES``, or ``KNOWN_UNMATCHED`` — or the transform
fails loudly, the same contract as crosswalk.json. State charter and
specialty schools (longer codes) have no F-33 counterpart and stay
parquet-only. Like F-33, everything here stays out of normalized.parquet
(state QBE aid is already counted as state Department of Education
spending).

Outputs:
- data/raw/gadoe_revenues_<year>.xlsx        export as downloaded (committed;
  the offline fallback when the portal is unreachable)
- data/processed/gadoe_revenues.parquet      one row per (system, fiscal
  year) across every year fetched over time
- data/processed/schools/gadoe.json          overlay for the years ahead of
  F-33, keyed by NCESID for the web (etl_schools.py leaves it in place)

Usage: etl_gadoe.py [gadoe_revenues]
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import duckdb
import openpyxl
import pandas as pd

import gadoe
import runlog
from fetching import call_with_retries

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
PARQUET_FILE = ROOT / "data" / "processed" / "gadoe_revenues.parquet"
SCHOOLS_DIR = ROOT / "data" / "processed" / "schools"
OVERLAY_FILE = SCHOOLS_DIR / "gadoe.json"
INDEX_FILE = SCHOOLS_DIR / "index.json"
SOURCE_ID = "gadoe_revenues"
REGULAR_CODE_LENGTH = 3
BASIS_NOTE = (
    "Operating revenues only: debt-service and capital-projects funds are "
    "excluded, so totals run below the all-funds Census F-33 figures."
)

WORKBOOK_COLUMNS = {
    "system_code": "System ID",
    "system_name": "System Name",
    "fte": "FTE Reported",
    "revenue_local": "Local Revenues",
    "per_fte_local": "Per FTE Local",
    "revenue_state": "State Revenues",
    "per_fte_state": "Per FTE State",
    "revenue_federal": "Federal Revenues",
    "per_fte_federal": "Per FTE Federal",
    "revenue_total": "Total Revenues",
    "per_fte_total": "Per FTE Total",
}
MONEY_FIELDS = [field for field in WORKBOOK_COLUMNS
                if field.startswith(("revenue_", "per_fte_"))]
STATE_TOTAL = "State Total"

MATCH_STOP_TOKENS = {"school", "district", "schools", "public", "independent",
                     "board", "of", "education", "system", "city"}
GADOE_ALIASES = {"Griffin-Spalding County": "griffin spalding"}
KNOWN_UNMATCHED = {
    "630": "Clay County",
    "731": "Taliaferro County",
    "752": "Webster County",
}


def match_key(name: str) -> str:
    tokens = re.sub(r"[^a-z ]", " ", name.lower()).split()
    return " ".join(token for token in tokens
                    if token not in MATCH_STOP_TOKENS)


def to_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").strip()
    return 0.0 if cleaned in ("", "-") else float(cleaned)


def clean_system_name(raw: str) -> str:
    return re.sub(r"\s*\*+$", "", str(raw).strip())


def numeric_fields(cells: tuple, columns: dict[str, int]) -> dict:
    values = {field: round(to_number(cells[columns[field]]), 2)
              for field in MONEY_FIELDS}
    values["fte"] = int(round(to_number(cells[columns["fte"]])))
    return values


def locate_columns(header: tuple) -> dict[str, int]:
    labels = [str(cell).strip() if cell is not None else "" for cell in header]
    columns = {
        field: next((position for position, label in enumerate(labels)
                     if label.startswith(prefix)), None)
        for field, prefix in WORKBOOK_COLUMNS.items()}
    missing = [WORKBOOK_COLUMNS[field]
               for field, position in columns.items() if position is None]
    if missing:
        raise SystemExit(
            f"Export lacks expected columns {missing} — layout may have changed.")
    return columns


def parse_workbook(path: Path, fiscal_year: int) -> tuple[list[dict], dict]:
    workbook = openpyxl.load_workbook(path, read_only=True)
    rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
    workbook.close()
    header_index = next(
        (position for position, row in enumerate(rows)
         if any(str(cell).strip() == "System ID" for cell in row if cell)),
        None)
    if header_index is None:
        raise SystemExit(
            f"{path.name} has no 'System ID' header row — layout may have changed.")
    preamble = " ".join(str(cell) for row in rows[:header_index]
                        for cell in row if cell)
    stated = re.search(r"Fiscal Year\s+(\d{4})", preamble)
    if not stated or int(stated.group(1)) != fiscal_year:
        raise SystemExit(
            f"{path.name} reports fiscal year "
            f"{stated.group(1) if stated else 'unknown'}, expected {fiscal_year}.")
    columns = locate_columns(rows[header_index])
    records: list[dict] = []
    statewide = None
    for cells in rows[header_index + 1:]:
        code = cells[columns["system_code"]]
        if code is None or not str(code).strip():
            continue
        if str(code).strip() == STATE_TOTAL:
            values = numeric_fields(cells, columns)
            statewide = {"fte": values["fte"],
                         **{field: values[field] for field in MONEY_FIELDS
                            if field.startswith("revenue_")}}
            break
        records.append({
            "fiscal_year": fiscal_year,
            "system_code": str(code).strip(),
            "system_name": clean_system_name(cells[columns["system_name"]]),
            **numeric_fields(cells, columns),
        })
    if not records or statewide is None:
        raise SystemExit(
            f"{path.name} yielded {len(records)} systems and "
            f"{'no' if statewide is None else 'a'} {STATE_TOTAL} row — "
            "layout may have changed.")
    return records, statewide


def load_index() -> dict:
    if not INDEX_FILE.exists():
        raise SystemExit(
            f"{runlog.display_path(INDEX_FILE)} is missing — run etl_schools.py "
            "first; the overlay is defined relative to the F-33 vintage.")
    return json.loads(INDEX_FILE.read_text())


def map_systems(records: list[dict], districts: list[dict]) -> dict[str, dict]:
    keyed = {}
    for district in districts:
        key = match_key(district["display_name"])
        if key in keyed:
            raise SystemExit(
                f"F-33 districts collide on match key {key!r}: "
                f"{keyed[key]['display_name']} / {district['display_name']}")
        keyed[key] = district
    mapping: dict[str, dict] = {}
    claimed: dict[str, str] = {}
    seen: set[str] = set()
    unmatched, stale_known = [], []
    for record in records:
        code, name = record["system_code"], record["system_name"]
        if len(code) != REGULAR_CODE_LENGTH or code in seen:
            continue
        seen.add(code)
        district = keyed.get(GADOE_ALIASES.get(name, match_key(name)))
        if district is None:
            if KNOWN_UNMATCHED.get(code) != name:
                unmatched.append(f"{code} {name}")
            continue
        if KNOWN_UNMATCHED.get(code) == name:
            stale_known.append(f"{code} {name}")
            continue
        if district["ncesid"] in claimed:
            raise SystemExit(
                f"GaDOE systems {claimed[district['ncesid']]} and {code} both "
                f"map to NCESID {district['ncesid']}.")
        claimed[district["ncesid"]] = code
        mapping[code] = district
    if unmatched or stale_known:
        raise SystemExit(
            "GaDOE system roster no longer matches the F-33 contract — "
            f"unmatched regular systems {unmatched or 'none'}, "
            f"KNOWN_UNMATCHED entries that now match {stale_known or 'none'}. "
            "Update GADOE_ALIASES / KNOWN_UNMATCHED deliberately.")
    return mapping


def ensure_raw(opener, year: int, year_value: str) -> Path:
    target = RAW_DIR / f"{SOURCE_ID}_{year}.xlsx"
    try:
        payload = call_with_retries(
            lambda: gadoe.export_revenues(opener, year_value, "EXCELOPENXML"),
            source=SOURCE_ID, description=f"revenues export FY{year}",
            attempts=3)
    except (Exception, SystemExit):
        if not target.exists():
            raise
        runlog.log_event("fallback", SOURCE_ID, path=runlog.display_path(target),
                         reason="download failed; using committed raw file")
        return target
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    target.write_bytes(payload)
    runlog.log_event("fetched", SOURCE_ID, url=gadoe.MENU_URL,
                     fiscal_year=year, bytes=target.stat().st_size)
    return target


def merge_parquet(fresh: pd.DataFrame) -> pd.DataFrame:
    if not PARQUET_FILE.exists():
        return fresh
    existing = duckdb.sql(f"FROM '{PARQUET_FILE}'").df()
    refreshed = set(fresh.fiscal_year.unique())
    kept = existing[~existing.fiscal_year.isin(refreshed)]
    return pd.concat([kept, fresh], ignore_index=True)


def write_parquet(frame: pd.DataFrame) -> None:
    PARQUET_FILE.parent.mkdir(parents=True, exist_ok=True)
    connection = duckdb.connect()
    connection.register("gadoe_revenues", frame)
    connection.execute(
        f"""
        COPY (SELECT * FROM gadoe_revenues ORDER BY fiscal_year, system_code)
        TO '{PARQUET_FILE}' (FORMAT PARQUET, COMPRESSION ZSTD)
        """
    )


def year_entry(record: dict) -> dict:
    sides = ("local", "state", "federal", "total")
    return {
        "fte": record["fte"],
        "revenue": {side: record[f"revenue_{side}"] for side in sides},
        "per_fte": {side: record[f"per_fte_{side}"] for side in sides},
    }


def statewide_entry(statewide: dict) -> dict:
    sides = ("local", "state", "federal", "total")
    fte = statewide["fte"]
    return {
        "fte": fte,
        "revenue": {side: statewide[f"revenue_{side}"] for side in sides},
        "per_fte": {side: round(statewide[f"revenue_{side}"] / fte, 2)
                    if fte else None for side in sides},
    }


def write_overlay(records: list[dict], statewide_by_year: dict[int, dict],
                  mapping: dict[str, dict], years: list[int],
                  f33_latest: int) -> int:
    districts: dict[str, dict] = {}
    for record in records:
        district = mapping.get(record["system_code"])
        if district is None:
            continue
        entry = districts.setdefault(district["ncesid"], {
            "system_code": record["system_code"],
            "system_name": record["system_name"],
            "slug": district["slug"],
            "years": {},
        })
        entry["years"][str(record["fiscal_year"])] = year_entry(record)
    document = {
        "source": SOURCE_ID,
        "report": gadoe.REVENUES_REPORT,
        "basis": BASIS_NOTE,
        "f33_latest_fiscal_year": f33_latest,
        "fiscal_years": years,
        "statewide": {str(year): statewide_entry(statewide)
                      for year, statewide in sorted(statewide_by_year.items())},
        "districts": dict(sorted(districts.items())),
    }
    SCHOOLS_DIR.mkdir(parents=True, exist_ok=True)
    OVERLAY_FILE.write_text(json.dumps(document, indent=1) + "\n")
    return len(districts)


def refresh() -> str:
    index = load_index()
    f33_latest = max(index["fiscal_years"])
    opener = gadoe.build_opener()
    years = call_with_retries(
        lambda: gadoe.year_values(gadoe.fetch_text(opener, gadoe.MENU_URL)),
        source=SOURCE_ID, description="fiscal-year menu", attempts=3)
    wanted = sorted(year for year in years if year > f33_latest)
    if not wanted:
        write_overlay([], {}, {}, [], f33_latest)
        return (f"F-33 already covers {f33_latest}; portal offers nothing "
                f"newer — wrote an empty overlay.")

    records: list[dict] = []
    statewide_by_year: dict[int, dict] = {}
    for year in wanted:
        path = ensure_raw(opener, year, years[year])
        year_records, statewide = parse_workbook(path, year)
        records += year_records
        statewide_by_year[year] = statewide
    mapping = map_systems(records, index["districts"])
    frame = merge_parquet(pd.DataFrame.from_records(records).assign(
        ncesid=lambda f: f.system_code.map(
            {code: district["ncesid"] for code, district in mapping.items()})))
    write_parquet(frame)
    mapped = write_overlay(records, statewide_by_year, mapping, wanted,
                           f33_latest)
    runlog.log_event("transformed", SOURCE_ID, records=len(records),
                     fiscal_years=wanted, districts_mapped=mapped)
    return (f"Wrote {len(records):,} system-year records "
            f"({', '.join(map(str, wanted))}; {mapped} districts mapped to "
            f"F-33) to {runlog.display_path(PARQUET_FILE)} and "
            f"{runlog.display_path(OVERLAY_FILE)}")


def main() -> int:
    requested = set(sys.argv[1:])
    if requested and SOURCE_ID not in requested:
        print(f"No GaDOE sources among {sorted(requested)}; nothing to do.")
        return 0
    try:
        summary = refresh()
    except (Exception, SystemExit) as exc:
        failures = runlog.record_outcome(SOURCE_ID, ok=False, error=str(exc))
        runlog.log_event("transform_failed", SOURCE_ID,
                         consecutive_failures=failures, error=str(exc)[:300])
        print(f"ERROR {SOURCE_ID}: {exc}", file=sys.stderr)
        return 1
    runlog.record_outcome(SOURCE_ID, ok=True)
    print(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
